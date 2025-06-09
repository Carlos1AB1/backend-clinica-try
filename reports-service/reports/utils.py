import os
import io
import csv
import json
import requests
from datetime import datetime, timedelta
from django.conf import settings
from django.http import HttpResponse
from django.core.cache import cache
from django.db import connection
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class ReportGenerator:
    """
    Generador de reportes en múltiples formatos
    """
    
    def __init__(self, template, parameters=None):
        self.template = template
        self.parameters = parameters or {}
        self.data = None
        
    def execute_query(self):
        """Ejecutar la consulta SQL del template"""
        try:
            # Procesar parámetros en la consulta
            query = self.template.sql_query
            for key, value in self.parameters.items():
                query = query.replace(f'${{{key}}}', str(value))
            
            # Ejecutar consulta
            with connection.cursor() as cursor:
                cursor.execute(query)
                columns = [col[0] for col in cursor.description]
                rows = cursor.fetchall()
                
            # Convertir a DataFrame para facilitar manipulación
            self.data = pd.DataFrame(rows, columns=columns)
            return True
            
        except Exception as e:
            print(f"Error ejecutando consulta: {e}")
            return False
    
    def fetch_external_data(self):
        """Obtener datos de otros microservicios"""
        data_sources = []
        
        # Definir qué datos necesitamos según la categoría del reporte
        if 'USUARIOS' in self.template.category:
            data_sources.append(self._fetch_users_data())
        if 'CITAS' in self.template.category:
            data_sources.append(self._fetch_appointments_data())
        if 'HISTORIAS' in self.template.category:
            data_sources.append(self._fetch_medical_records_data())
        if 'RECETAS' in self.template.category:
            data_sources.append(self._fetch_prescriptions_data())
        
        return data_sources
    
    def _fetch_users_data(self):
        """Obtener datos del microservicio de usuarios"""
        try:
            response = requests.get(
                f"{settings.USERS_SERVICE_URL}/api/v1/owners/",
                params=self.parameters,
                timeout=30
            )
            if response.status_code == 200:
                return response.json()
        except requests.RequestException:
            pass
        return []
    
    def _fetch_appointments_data(self):
        """Obtener datos del microservicio de citas"""
        try:
            response = requests.get(
                f"{settings.APPOINTMENTS_SERVICE_URL}/api/v1/appointments/",
                params=self.parameters,
                timeout=30
            )
            if response.status_code == 200:
                return response.json()
        except requests.RequestException:
            pass
        return []
    
    def _fetch_medical_records_data(self):
        """Obtener datos del microservicio de historias clínicas"""
        try:
            response = requests.get(
                f"{settings.MEDICAL_RECORDS_SERVICE_URL}/api/v1/medical-records/",
                params=self.parameters,
                timeout=30
            )
            if response.status_code == 200:
                return response.json()
        except requests.RequestException:
            pass
        return []
    
    def _fetch_prescriptions_data(self):
        """Obtener datos del microservicio de recetas"""
        try:
            response = requests.get(
                f"{settings.PRESCRIPTIONS_SERVICE_URL}/api/v1/prescriptions/prescriptions/",
                params=self.parameters,
                timeout=30
            )
            if response.status_code == 200:
                return response.json()
        except requests.RequestException:
            pass
        return []
    
    def generate_pdf(self, filename):
        """Generar reporte en PDF"""
        if self.data is None or self.data.empty:
            return None
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Título del reporte
        title = Paragraph(self.template.name, title_style)
        elements.append(title)
        
        # Información del reporte
        info_text = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>"
        info_text += f"Total de registros: {len(self.data)}"
        info = Paragraph(info_text, subtitle_style)
        elements.append(info)
        elements.append(Spacer(1, 20))
        
        # Preparar datos para la tabla
        table_data = []
        headers = [str(col) for col in self.data.columns]
        table_data.append(headers)
        
        # Agregar filas (limitar a 1000 para PDFs)
        for _, row in self.data.head(1000).iterrows():
            table_data.append([str(cell) for cell in row])
        
        # Crear tabla
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Si hay más de 1000 registros, agregar nota
        if len(self.data) > 1000:
            note = Paragraph(
                f"<b>Nota:</b> Se muestran los primeros 1000 registros de {len(self.data)} totales.",
                styles['Normal']
            )
            elements.append(Spacer(1, 20))
            elements.append(note)
        
        # Construir PDF
        doc.build(elements)
        
        # Guardar archivo
        buffer.seek(0)
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        with open(file_path, 'wb') as f:
            f.write(buffer.getvalue())
        
        return file_path
    
    def generate_excel(self, filename):
        """Generar reporte en Excel"""
        if self.data is None or self.data.empty:
            return None
        
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Título del reporte
        ws.merge_cells('A1:' + chr(65 + len(self.data.columns) - 1) + '1')
        ws['A1'] = self.template.name
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Información del reporte
        ws.merge_cells('A2:' + chr(65 + len(self.data.columns) - 1) + '2')
        ws['A2'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: {len(self.data)} registros"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Headers
        start_row = 4
        for col_num, column_title in enumerate(self.data.columns, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for row_num, (_, row) in enumerate(self.data.iterrows(), start_row + 1):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        wb.save(file_path)
        return file_path
    
    def generate_csv(self, filename):
        """Generar reporte en CSV"""
        if self.data is None or self.data.empty:
            return None
        
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Guardar CSV
        self.data.to_csv(file_path, index=False, encoding='utf-8-sig')
        return file_path
    
    def generate_html_response(self):
        """Generar reporte como respuesta HTML"""
        if self.data is None or self.data.empty:
            return None
        
        # Convertir DataFrame a HTML
        html_table = self.data.to_html(classes='table table-striped table-bordered', 
                                      table_id='reportTable', 
                                      escape=False)
        
        # Template HTML completo
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{self.template.name}</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/js/bootstrap.bundle.min.js"></script>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .report-header {{ text-align: center; margin-bottom: 30px; }}
                .report-info {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                .table {{ font-size: 12px; }}
                .table th {{ background-color: #6c757d; color: white; }}
                .export-buttons {{ margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <div class="container-fluid">
                <div class="report-header">
                    <h1>{self.template.name}</h1>
                    <p class="text-muted">{self.template.description}</p>
                </div>
                
                <div class="report-info">
                    <div class="row">
                        <div class="col-md-4">
                            <strong>Generado:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M')}
                        </div>
                        <div class="col-md-4">
                            <strong>Total registros:</strong> {len(self.data)}
                        </div>
                        <div class="col-md-4">
                            <strong>Categoría:</strong> {self.template.category}
                        </div>
                    </div>
                </div>
                
                <div class="table-responsive">
                    {html_table}
                </div>
                
                <div class="mt-4 text-center">
                    <p class="text-muted">
                        Reporte generado por el Sistema de Clínica Veterinaria
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return HttpResponse(html_content, content_type='text/html')

def cache_report_data(key, data, timeout=3600):
    """Cachear datos de reporte"""
    cache.set(key, data, timeout)

def get_cached_report_data(key):
    """Obtener datos de reporte desde cache"""
    return cache.get(key)

def generate_cache_key(template_id, parameters):
    """Generar clave de cache para reporte"""
    param_str = json.dumps(parameters, sort_keys=True)
    return f"report_{template_id}_{hash(param_str)}" 